import pandas as pd
import os
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def read_xls_excel(url,index):
    '''
    读取xls格式文件
    参数：
        url:文件路径
        index：工作表序号（第几个工作表，传入参数从1开始数）
    返回：
        data:表格中的数据
    '''
    # 打开指定的工作簿
    workbook = xlrd.open_workbook(url)
    # 获取工作簿中的所有表格
    sheets = workbook.sheet_names()
    # 获取工作簿中所有表格中的的第 index 个表格
    worksheet = workbook.sheet_by_name(sheets[index-1])
    # 定义列表存储表格数据
    data = []
    # 遍历每一行数据
    for i in range(0, worksheet.nrows):
        # 定义表格存储每一行数据
        da = []
        # 遍历每一列数据
        for j in range(0, worksheet.ncols):
            # 将行数据存储到da列表
            da.append(worksheet.cell_value(i, j))
        # 存储每一行数据
        data.append(da)
    # 返回数据
    # print(data)
    return data

def write_file(path, save):
    xls = []
    count = 0
    for file in os.listdir(path):
        ## 后缀
        suff = os.path.splitext(file)[1]

        ## pdf转docx
        if suff == '.xls' or suff == '.xlsx':
            print(file)
            if count == 0:
                data = read_xls_excel(path+file, 1)
                xls += data
                count += 1
            else: 
                data = read_xls_excel(path+file, 1)
                xls += data[1:]
            
    output = Workbook()
    wb = output.active

    for i in range(len(xls)):
        wb.append(xls[i])
    output.save(save + "所有数据.xlsx")    


if __name__ == "__main__":
    read_path = r"C:\\Users\\Administrator\\Desktop\\code\\"
    save_path = r"C:\\Users\\Administrator\\Desktop\\"
    write_file(read_path, save_path)

    # 读取Excel文件
    df = pd.read_excel(save_path + '所有数据.xlsx')

    # 获取重复项中折扣最低的一组数据
    df_duplicates = df.sort_values('折扣').drop_duplicates(['条码', '定价'], keep='first')

    # 删除重复项
    df_final = df_duplicates.drop_duplicates(['条码', '定价'], keep=False)

    # 排序
    df_final = df_final.sort_index(ascending=True, ignore_index=True)
    
    
    # 根据条码进行排序
    sorted_data = df.sort_values(['条码', '定价'])
    # 提取重复的数据
    duplicated_data = sorted_data[sorted_data.duplicated(['条码', '定价'], keep=False)]
    # 输出重复数据到新的XLS文件
    duplicated_data.to_excel(save_path + '重复数据.xlsx', index=False)

    # 将处理后的数据写入新的Excel文件
    df_final.to_excel(save_path + '最终数据.xlsx')
